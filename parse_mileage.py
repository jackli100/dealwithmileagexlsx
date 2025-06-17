import sys
import re
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

def parse_value(val):
    if val is None:
        return None, None
    s = str(val).strip()
    # case a: DKxxx+yyy-DKxxx+zzz
    m = re.fullmatch(r"DK(\d+)\+([\d.]+)\s*-\s*DK(\d+)\+([\d.]+)", s, re.IGNORECASE)
    if m:
        start = float(m.group(1)) * 1000 + float(m.group(2))
        end = float(m.group(3)) * 1000 + float(m.group(4))
        return start, end
    # case b: DKxxx+yyy(.zzz)
    m = re.fullmatch(r"DK(\d+)\+([\d.]+)", s, re.IGNORECASE)
    if m:
        start = float(m.group(1)) * 1000 + float(m.group(2))
        return start, None
    # case c: pure digits, length>9 and even
    if re.fullmatch(r"\d+", s) and len(s) > 9 and len(s) % 2 == 0:
        mid = len(s) // 2
        start = float(s[:mid])
        end = float(s[mid:])
        return start, end
    # case d: pure digits with optional decimal
    if re.fullmatch(r"\d+(?:\.\d+)?", s):
        return float(s), None
    # case e: keep as is
    return s, None

def main(path, column_letter):
    wb = load_workbook(path)
    ws = wb.active
    # remove merged cells to avoid read-only errors
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    idx = column_index_from_string(column_letter)

    # read original values before inserting a new column
    values = [ws.cell(row=i, column=idx).value for i in range(1, ws.max_row + 1)]

    ws.insert_cols(idx + 1)
    ws.cell(row=1, column=idx).value = "起始里程"
    ws.cell(row=1, column=idx + 1).value = "结束里程"

    for row in range(2, len(values) + 1):
        val = values[row - 1]
        start, end = parse_value(val)
        ws.cell(row=row, column=idx).value = start
        ws.cell(row=row, column=idx + 1).value = end

    out_path = path.rsplit('.', 1)[0] + '_里程解析后.xlsx'
    wb.save(out_path)

if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("Usage: python parse_mileage.py <xlsx_path> <column_letter>")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
